import os

import openpyxl
import weasyprint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html


def format_areas(filepath: str, header_color: str = "92d050"):
    # Load the Excel file to apply formatting
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Set the alignment to center for all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply green background to the header row
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted Excel file
    wb.save(filepath)
    return filepath


def format_service_charge(filepath: str, master_project_en: str, header_color: str = "92d050"):
    """
    Formats an Excel file by adding specific headers and applying styles.

    Parameters:
    - filepath (str): Path to the Excel file to format.
    - b1_value (str): Value to write in cell B1.
    - header_color (str): Hex color code for the header background. Default is "92d050".

    Returns:
    - str: The filepath of the formatted Excel file.
    """
    # Load the Excel workbook and select the active worksheet
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 1. Insert two new rows at the top to make space for headers
    ws.insert_rows(1, amount=2)

    # 2. Write "master_project_en" in cell A1
    ws['A1'] = "master_project_en"
    
    # Make it bold
    ws['A1'].font = openpyxl.styles.Font(bold=True)

    # 3. Write the provided b1_value in cell B1
    ws['B1'] = master_project_en

    # 4. Write "Service charges (AED per sqft/year)" across cells D2 to H2
    # Merge cells D2 to H2 for a unified header
    ws.merge_cells('C2:G2')
    ws['C2'] = "Service charges (AED per sqft/year)"
    
    # Make it bold
    ws['C2'].font = openpyxl.styles.Font(bold=True)

    # Center the merged cell both horizontally and vertically
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

    ws['A1'].fill = header_fill
    for cell in ["A3", "B3", "C3", "D3", "E3", "F3", "G3"]:
        ws[cell].fill = header_fill

    # Auto-adjust column widths based on the maximum length in each column
    for col in [ws[get_column_letter(i)] for i in range(1, 2+1)]:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value else ""
                cell_length = len(cell_value)
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2  # Add extra padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted Excel file
    wb.save(filepath)
    return filepath


def format_query_files(filepath: str, header_color: str = "92d050"):
    # Load the Excel file to apply formatting
    wb = openpyxl.load_workbook(filepath)


def format_calculator_result(transaction, filename: str):
    """ Format the calculator result into an Excel file """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define styles
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="fef1cc", end_color="fef1cc", fill_type="solid")
    green_fill = PatternFill(start_color="d2f1da", end_color="d2f1da", fill_type="solid")

    # Add project name and calculation date
    ws.append(["PROJECT NAME", transaction.projectName])
    ws["A1"].font = bold_font

    # Add calculation date
    ws.append(["CALCULATION DATE", transaction.calculationDate])

    # Add blank row
    ws.append([])

    # Add price net
    ws.append(["PRICE NET:", f"AED {transaction.salePrice}"])
    ws["A4"].font = bold_font
    ws["A4"].fill = yellow_fill
    ws["B4"].fill = yellow_fill

    # Add blank row
    ws.append([])

    # Add payment on transfer
    ws.append(["PAYMENT ON TRANSFER:", f"AED {transaction.paymentTransfer}"])
    ws["A6"].font = bold_font
    ws["A6"].fill = green_fill
    ws["B6"].fill = green_fill

    # Add payment details
    ws.append(["1. Payment to Seller", f"AED {transaction.paymentSeller}"])
    ws.append([f"2. DLD fee 4% + {transaction.constructionFee} AED", f"AED {transaction.dldFee}"])
    ws.append(["3. DLD Registration Trustee fee + 5%VAT", f"AED {transaction.registrationTrusteeFee}"])
    ws.append(["4. Buyer's agent commission 2% + 5%VAT", f"AED {transaction.agentCommission}"])

    if transaction.managersChequePercent:
        manager_cheque_string = f"Manager cheque {transaction.managersChequePercent}%"
        parties = [
            "Seller" if transaction.sellerCheque else None,
            "DLD" if transaction.dldCheque else None,
            "Comission" if transaction.commissionCheque else None
        ]
        parties = [party for party in parties if party]
        if len(parties) > 1:
            manager_cheque_string += f" ({', '.join(parties)})"
        if len(parties) == 1:
            manager_cheque_string += f" ({parties[0]})"

        ws.append([manager_cheque_string, f"AED {transaction.managersChequeAmount}"])
    # Add blank row
    ws.append([])

    if transaction.mortgagePaymentsAmounts:
        # Add payment plan to developer
        ws.append(["PAYMENT PLAN TO DEVELOPER:", f"AED {transaction.paymentPlan}"])
        ws["A13"].font = bold_font
        ws["A13"].fill = green_fill
        ws["B13"].fill = green_fill

        # Add mortgage payments
        if transaction.mortgagePaymentsDates:
            for payment, date in zip(
                    transaction.mortgagePaymentsAmounts,
                    transaction.mortgagePaymentsDates
                ):
                ws.append([date, f"AED {payment}"])
        elif transaction.mortgagePaymentsPercents:
            for payment, percent in zip(
                    transaction.mortgagePaymentsAmounts,
                    transaction.mortgagePaymentsPercents
                ):
                ws.append([f"{percent}% construction", f"AED {payment}"])
        else:
            for payment in transaction.mortgagePaymentsAmounts:
                ws.append(["", f"AED {payment}"])

    # Add blank row
    ws.append([])

    # Add price including all fees
    ws.append(["PRICE INCL. ALL FEES:", f"AED {transaction.totalPrice}"])


    # Adjust column widths
    for col in range(1, 3):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:  # noqa: E722
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)


def to_pdf(filepath: str):
    """ Convert an Excel file to a PDF file """
    dir = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    basename, ext = os.path.splitext(filename)
    output_filepath = f'{dir}/{basename}.pdf'

    if ext != '.xlsx':
        raise ValueError("Input file must be an Excel file.")

    xlsx2html(filepath, f'{dir}/{basename}.html')
    pdf = weasyprint.HTML(f'{dir}/{basename}.html').write_pdf()
    with open(output_filepath, 'wb') as f:
        f.write(pdf)
    return output_filepath
