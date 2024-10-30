import logging.config
import os

import openpyxl
import pandas as pd
from omegaconf import OmegaConf
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from real_estate_telegram_bot.api.users import check_user_in_channel_sync
from real_estate_telegram_bot.db import crud
from telebot.types import InlineKeyboardButton, InlineKeyboardMarkup

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

config = OmegaConf.load("./src/real_estate_telegram_bot/conf/config.yaml")
strings = config.strings

# Custom sorting key
def custom_sort_key(val):
    if isinstance(val, str):
        return (0, val)  # Strings get a tuple with 0 as the first element
    else:
        return (1, val)  # Integers get a tuple with 1 as the first element
    
def create_main_menu_markup(strings):
    menu_markup = InlineKeyboardMarkup(row_width=1)
    menu_markup.add(
        InlineKeyboardButton(strings.menu.query, callback_data="_query"),
        InlineKeyboardButton(strings.menu.files, callback_data="_files"),
        InlineKeyboardButton(strings.menu.areas, callback_data="_areas"),
        InlineKeyboardButton(strings.menu.area_names, callback_data="_area_names"),
        InlineKeyboardButton(strings.menu.useful_links, callback_data="_useful_links"),
        InlineKeyboardButton(strings.menu.support, callback_data="_support"),
        InlineKeyboardButton(strings.menu.language, callback_data="_language"),
    )
    return menu_markup

def create_lang_menu_markup(strings):
    lang_menu_markup = InlineKeyboardMarkup(row_width=1)
    lang_menu_markup.add(
        InlineKeyboardButton(strings.language_en, callback_data="_en"),
        InlineKeyboardButton(strings.language_ru, callback_data="_ru")
    )
    return lang_menu_markup

def create_main_menu_button(strings):
    main_menu_button = InlineKeyboardMarkup(row_width=1)
    main_menu_button.add(InlineKeyboardButton(strings.main_menu, callback_data="_main_menu"))
    return main_menu_button

def create_create_query_menu(strings):
    query_menu = InlineKeyboardMarkup(row_width=2)
    query_menu.add(InlineKeyboardButton(strings.menu.query, callback_data="_query"))
    query_menu.add(InlineKeyboardButton(strings.main_menu, callback_data="_main_menu"))
    return query_menu

def create_areas_menu_markup(strings: dict, lang: str):
    areas_menu_markup = InlineKeyboardMarkup(row_width=2)
    areas_menu_markup.add(
        InlineKeyboardButton(strings.areas.al_furjan, callback_data="_al_furjan"),
        InlineKeyboardButton(strings.areas.arjan, callback_data="_arjan"),
        InlineKeyboardButton(strings.areas.beachfront, callback_data="_beachfront"),
        InlineKeyboardButton(strings.areas.bluewaters, callback_data="_bluewaters"),
        InlineKeyboardButton(strings.areas.business_bay, callback_data="_business_bay"),
        InlineKeyboardButton(strings.areas.city_walk, callback_data="_city_walk"),
        InlineKeyboardButton(strings.areas.creek_harbour, callback_data="_creek_harbour"),
        InlineKeyboardButton(strings.areas.downtown, callback_data="_downtown"),
        InlineKeyboardButton(strings.areas.dubai_hills, callback_data="_dubai_hills"),
        InlineKeyboardButton(strings.areas.dubai_islands, callback_data="_dubai_islands"),
        InlineKeyboardButton(strings.areas.dubai_marina, callback_data="_dubai_marina"),
        InlineKeyboardButton(strings.areas.dubai_maritime_city, callback_data="_dubai_maritime_city"),
        InlineKeyboardButton(strings.areas.jlt, callback_data="_jlt"),
        InlineKeyboardButton(strings.areas.jvc, callback_data="_jvc"),
        InlineKeyboardButton(strings.areas.jvt, callback_data="_jvt"),
        InlineKeyboardButton(strings.areas.jbr, callback_data="_jbr"),
        InlineKeyboardButton(strings.areas.la_mer, callback_data="_la_mer"),
        InlineKeyboardButton(strings.areas.mina_rashid, callback_data="_mina_rashid"),
        InlineKeyboardButton(strings.areas.palm_jumeirah, callback_data="_palm_jumeirah"),
        InlineKeyboardButton(strings.areas.sobha_hartland, callback_data="_sobha_hartland"),
        InlineKeyboardButton(strings[lang].query.enter_own_area, callback_data="_enter_own_area")
    )
    return areas_menu_markup

def register_handlers(bot):
    @bot.message_handler(commands=["start", "menu"])
    def menu_menu_command(message):

        user_id = message.from_user.id
        username = message.from_user.username

        # Check if user is in the channel
        if check_user_in_channel_sync(config.channel_name, username) is False:
            bot.send_message(
                message.chat.id,
                f"You need to join the channel @{config.channel_name} to use the bot."
            )
            return
        crud.upsert_user(user_id, username)

        user = crud.read_user(user_id)
        lang = user.language
        logger.info({"user_id": message.from_user.id, "message": message.text})

        bot.send_message(
            message.chat.id, strings[lang].start,
            reply_markup=create_main_menu_markup(strings[lang])
        )

    @bot.callback_query_handler(func=lambda call: call.data == "_main_menu")
    def main_menu_callback(call):
        user_id = call.from_user.id
        username = call.from_user.username

        user = crud.read_user(user_id)
        if not user:
            user = crud.upsert_user(user_id, username)

        lang = user.language
        logger.info({"user_id": user_id, "message": call.data})

        bot.send_message(
            call.message.chat.id, strings[lang].start,
            reply_markup=create_main_menu_markup(strings[lang])
        )

    # Language selection
    @bot.callback_query_handler(func=lambda call: call.data == "_language")
    def language(call):
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language

        logger.info({"user_id": call.from_user.id, "message": call.data})

        lang_menu_markup = create_lang_menu_markup(strings[lang])
        bot.send_message(call.message.chat.id, "Language", reply_markup=lang_menu_markup)

    @bot.callback_query_handler(func=lambda call: call.data == "_ru")
    def language_ru(call):
        logger.info({"user_id": call.from_user.id, "message": call.data})
        user_id = call.from_user.id
        username = call.from_user.username
        lang = "ru"
        crud.update_user_language(user_id, new_language=lang)

        bot.send_message(
            call.message.chat.id, strings[lang].language_selected,
            reply_markup=create_main_menu_button(strings[lang])
        )

    @bot.callback_query_handler(func=lambda call: call.data == "_en")
    def language_en(call):
        logger.info({"user_id": call.from_user.id, "message": call.data})
        user_id = call.from_user.id
        username = call.from_user.username

        lang = "en"
        crud.update_user_language(user_id, new_language=lang)

        bot.send_message(
            call.message.chat.id, strings[lang].language_selected,
            reply_markup=create_main_menu_button(strings[lang])
        )


    # Useful links
    @bot.callback_query_handler(func=lambda call: call.data == "_useful_links")
    def useful_links(call):
        logger.info({"user_id": call.from_user.id, "message": call.data})
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language

        bot.send_message(call.message.chat.id, strings[lang].useful_links,
            reply_markup=create_main_menu_button(strings[lang])
        )

    # Help
    @bot.callback_query_handler(func=lambda call: call.data == "_support")
    def support(call):
        logger.info({"user_id": call.from_user.id, "message": call.data})
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language

        bot.send_message(call.message.chat.id, strings[lang].support,
            reply_markup=create_main_menu_button(strings[lang])
        )

    # Query
    @bot.callback_query_handler(func=lambda call: call.data == "_query")
    def query(call):
        logger.info({"user_id": call.from_user.id, "message": call.data})
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language
        bot.send_message(
            call.message.chat.id, strings[lang].query.ask_name,
            reply_markup=create_main_menu_button(strings[lang])
        )

    # Query area names table
    @bot.callback_query_handler(func=lambda call: call.data == "_area_names")
    def get_area_names_table(call):
        logger.info({"user_id": call.from_user.id, "message": call.data})
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language

        # Send the downloaded file to the user
        with open("./data/dubai_area_names.xlsx", 'rb') as file:
            bot.send_document(user_id, file, reply_markup=create_main_menu_button(strings[lang])
        )

    # Areas menu handler
    @bot.callback_query_handler(func=lambda call: call.data == "_areas")
    def areas_menu_callback(call):
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language

        logger.info({"user_id": user_id, "message": call.data})

        bot.send_message(
            call.message.chat.id, strings[lang].menu.select_area,
            reply_markup=create_areas_menu_markup(strings, lang)
        )

    # Specific area handlers (e.g., for "Dubai Marina")
    @bot.callback_query_handler(func=lambda call: call.data in [
    "_dubai_marina", "_business_bay", "_downtown", "_creek_harbour", "_sobha_hartland", "_city_walk",
    "_al_furjan", "_arjan", "_beachfront", "_bluewaters", "_dubai_hills", "_dubai_islands",
    "_dubai_maritime_city", "_jlt", "_jvc", "_jvt", "_jbr", "_la_mer", "_mina_rashid",
    "_palm_jumeirah"])
    def area_callback(call):
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language

        if call.data == "_jlt":
            area_name = "Jumeirah Lakes Towers"
        elif call.data == "_jvc":
            area_name = "Jumeirah Village Circle"
        elif call.data == "_jvt":
            area_name = "Jumeirah Village Triangle"
        elif call.data == "_jbr":
            area_name = "Jumeirah Beach Residence"

        area_name = call.data[1:].replace("_", " ").title()  # Format the area name from the callback data
        building_data = crud.get_buildings_by_area(area_name)

        # Create a pandas DataFrame from the sorted data
        df = pd.DataFrame(building_data)
 
        df = df.sort_values(by=['How old is the building (years)'], key=lambda col: col.map(custom_sort_key))

        # Define the filename for the Excel file
        filename = f"{area_name}_buildings.xlsx"

        # check if tmp exists
        if not os.path.exists("./tmp"):
            os.makedirs("./tmp")
        filepath = os.path.join("./tmp", filename)  # Save it to a temporary directory


        # Save the DataFrame to a local Excel file
        df.to_excel(filepath, index=False)

        format_excel_file(filepath)

        # Send the Excel file to the user
        with open(filepath, 'rb') as file:
            bot.send_document(
                call.message.chat.id, file,
                reply_markup=create_main_menu_button(strings[lang])
            )

        # Delete the Excel file after sending it
        os.remove(filepath)

    # Handle the option where the user enters their own area name
    @bot.callback_query_handler(func=lambda call: call.data == "_enter_own_area")
    def enter_own_area_callback(call):
        user_id = call.from_user.id
        user = crud.read_user(user_id)
        lang = user.language
        msg = bot.send_message(call.message.chat.id, strings[lang].query.enter_own_area)
        bot.register_next_step_handler(msg, process_area_name)

    def process_area_name(message):
        user_id = message.from_user.id
        area_name = message.text
        user = crud.read_user(user_id)
        lang = user.language
        building_data = crud.get_buildings_by_area(area_name)
        if building_data:
            # Create a pandas DataFrame from the sorted data
            df = pd.DataFrame(building_data)

            # sort by `Construction end date` keeping strings at the top
            df = df.sort_values(by=['How old is the building (years)'], key=lambda col: col.map(custom_sort_key))


            # Define the filename for the Excel file
            filename = f"{area_name}_buildings.xlsx"
            filepath = os.path.join("./tmp", filename)  # Save it to a temporary directory

            # Save the DataFrame to a local Excel file
            df.to_excel(filepath, index=False)

            # Format the Excel file
            format_excel_file(filepath)

            # Send the formatted Excel file to the user
            with open(filepath, 'rb') as file:
                bot.send_document(user_id, file, reply_markup=create_main_menu_button(strings[lang]))

            # Delete the Excel file after sending it
            os.remove(filepath)
        else:
            bot.send_message(
                message.chat.id,
                strings[lang].area_query.result_negative,
                reply_markup=create_create_query_menu(strings[lang])
            )

def format_excel_file(filepath):
    # Load the Excel file to apply formatting
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Set the alignment to center for all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted Excel file
    wb.save(filepath)
    return filepath
