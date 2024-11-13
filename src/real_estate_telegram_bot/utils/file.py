import openpyxl
from openpyxl.styles import Alignment, PatternFill


def format_excel_file(filepath: str, header_color: str = "92d050"):
    # Load the Excel file to apply formatting
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Set the alignment to center for all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply green background to the header row
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted Excel file
    wb.save(filepath)
    return filepath
